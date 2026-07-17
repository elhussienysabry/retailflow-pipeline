"""
RetailFlow Pipeline — Analytics Excel Exporter
===============================================

Connects to the PostgreSQL warehouse, runs 4 analytics queries, and exports
the results to a professionally styled Excel workbook with separate sheets.

Queries:
    - Top Customers:   Top 10 customers by total net revenue
    - Monthly Sales:    Month-over-month revenue and order counts
    - Category Perf:    Revenue and units sold per product category
    - Cohort Analysis:  Customer retention by first-purchase month

Usage:
    python -m src.exports.excel_exporter

Environment variables:
    DB_HOST / POSTGRES_HOST
    DB_PORT / POSTGRES_PORT
    DB_NAME / POSTGRES_DB
    DB_USER / POSTGRES_USER
    DB_PASSWORD / POSTGRES_PASSWORD
"""

import logging
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
OUTPUT_DIR = PROJECT_ROOT / "outputs"

SHEET_NAMES: List[str] = [
    "Top Customers",
    "Monthly Sales",
    "Category Performance",
    "Cohort Analysis",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
CURRENCY_FORMAT = "#,##0.00"
DATE_FORMAT = "YYYY-MM-DD"


def _get_env(key: str, fallback: str) -> str:
    """Read an environment variable with a fallback key.

    Args:
        key: Primary env var name.
        fallback: Fallback env var name.

    Returns:
        The env var value, or an empty string if neither is set.
    """
    return os.getenv(key) or os.getenv(fallback) or ""


def get_engine() -> Engine:
    """Create a SQLAlchemy engine from environment variables.

    Supports DB_HOST/POSTGRES_HOST, DB_PORT/POSTGRES_PORT, etc.

    Returns:
        A SQLAlchemy Engine connected to PostgreSQL.

    Raises:
        ValueError: If required connection variables are missing.
    """
    host = _get_env("DB_HOST", "POSTGRES_HOST") or "localhost"
    port = _get_env("DB_PORT", "POSTGRES_PORT") or "5432"
    database = _get_env("DB_NAME", "POSTGRES_DB") or "retailflow"
    user = _get_env("DB_USER", "POSTGRES_USER") or "retailflow_user"
    password = _get_env("DB_PASSWORD", "POSTGRES_PASSWORD") or "retailflow_pass"

    conn_str = f"postgresql+psycopg2://{user}:{password}@{host}:{port}/{database}"
    engine = create_engine(conn_str, echo=False)
    logger.info("Connected to PostgreSQL at %s:%s/%s", host, port, database)
    return engine


ANALYTICS_QUERIES: Dict[str, str] = {
    "Top Customers": """
        SELECT
            c.first_name,
            c.last_name,
            c.email,
            c.country,
            c.city,
            ROUND(SUM(f.net_revenue_dollars), 2) AS total_net_revenue
        FROM marts.dim_customers AS c
        INNER JOIN marts.fct_orders AS f
            ON c.customer_id = f.customer_id
        WHERE f.status = 'completed'
        GROUP BY c.customer_id, c.first_name, c.last_name,
                 c.email, c.country, c.city
        ORDER BY total_net_revenue DESC
        LIMIT 10
    """,
    "Monthly Sales": """
        SELECT
            DATE_TRUNC('month', order_date)::DATE AS month,
            COUNT(DISTINCT order_id) AS total_orders,
            ROUND(SUM(net_revenue_dollars), 2) AS net_revenue,
            ROUND(
                (SUM(net_revenue_dollars)
                 - LAG(SUM(net_revenue_dollars))
                   OVER (ORDER BY DATE_TRUNC('month', order_date)))
                / NULLIF(LAG(SUM(net_revenue_dollars))
                         OVER (ORDER BY DATE_TRUNC('month', order_date)), 0)
                * 100, 2
            ) AS month_over_month_growth_pct
        FROM marts.fct_orders
        WHERE status = 'completed'
        GROUP BY DATE_TRUNC('month', order_date)
        ORDER BY month
    """,
    "Category Performance": """
        SELECT
            p.category,
            COUNT(DISTINCT f.order_id) AS total_orders,
            SUM(f.quantity) AS total_units_sold,
            ROUND(SUM(f.net_revenue_dollars), 2) AS total_net_revenue,
            ROUND(AVG(f.discount_pct), 2) AS avg_discount_pct,
            ROUND(
                SUM(f.net_revenue_dollars)
                / SUM(SUM(f.net_revenue_dollars)) OVER ()
                * 100, 2
            ) AS revenue_share_pct
        FROM marts.dim_products AS p
        INNER JOIN marts.fct_orders AS f
            ON p.product_id = f.product_id
        WHERE f.status = 'completed'
        GROUP BY p.category
        ORDER BY total_net_revenue DESC
    """,
    "Cohort Analysis": """
        WITH customer_first_order AS (
            SELECT
                customer_id,
                DATE_TRUNC('month', MIN(order_date))::DATE AS cohort_month
            FROM marts.fct_orders
            WHERE status = 'completed'
            GROUP BY customer_id
        ),
        customer_monthly AS (
            SELECT
                cfo.cohort_month,
                EXTRACT(MONTH FROM AGE(f.order_date, cfo.cohort_month))::INTEGER
                    AS cohort_index,
                COUNT(DISTINCT f.customer_id) AS active_customers,
                ROUND(SUM(f.net_revenue_dollars), 2) AS total_revenue,
                ROUND(AVG(f.net_revenue_dollars), 2) AS avg_revenue_per_customer
            FROM marts.fct_orders AS f
            INNER JOIN customer_first_order AS cfo
                ON f.customer_id = cfo.customer_id
            WHERE f.status = 'completed'
            GROUP BY cfo.cohort_month, cohort_index
        )
        SELECT
            cohort_month,
            cohort_index,
            active_customers,
            total_revenue,
            avg_revenue_per_customer
        FROM customer_monthly
        ORDER BY cohort_month, cohort_index
    """,
}


def fetch_data(engine: Engine) -> Dict[str, pd.DataFrame]:
    """Execute all analytics queries and return DataFrames.

    Args:
        engine: SQLAlchemy Engine connected to PostgreSQL.

    Returns:
        Dict mapping sheet name to query result DataFrame.
    """
    results: Dict[str, pd.DataFrame] = {}
    for sheet_name, sql in ANALYTICS_QUERIES.items():
        logger.info("Executing query: %s", sheet_name)
        try:
            df = pd.read_sql(text(sql), engine)
            results[sheet_name] = df
            logger.info("  -> %d rows retrieved", len(df))
        except Exception as exc:
            logger.error("Query '%s' failed: %s", sheet_name, exc)
            results[sheet_name] = pd.DataFrame()
    return results


def _auto_fit_columns(ws: Any, df: pd.DataFrame) -> None:
    """Adjust column widths to fit content (header + data).

    Args:
        ws: openpyxl worksheet.
        df: DataFrame whose data populates the sheet.
    """
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for value in df[col_name].head(50):
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _style_sheet(ws: Any, df: pd.DataFrame) -> None:
    """Apply professional styling to a worksheet.

    Args:
        ws: openpyxl worksheet.
        df: DataFrame whose data populates the sheet.
    """
    header = list(ws.iter_rows(min_row=1, max_row=1))[0]
    for cell in header:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal="right")

    _auto_fit_columns(ws, df)


def _format_currency_columns(ws: Any, df: pd.DataFrame) -> None:
    """Apply currency number format to columns containing revenue or price.

    Args:
        ws: openpyxl worksheet.
        df: DataFrame whose data populates the sheet.
    """
    currency_keywords = ["revenue", "price", "amount", "spend"]
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_lower = col_name.lower().replace("_", " ")
        if any(kw in col_lower for kw in currency_keywords):
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell.number_format = CURRENCY_FORMAT


def build_workbook(data: Dict[str, pd.DataFrame]) -> Workbook:
    """Build a styled Excel workbook with one sheet per query.

    Args:
        data: Dict mapping sheet name to query result DataFrame.

    Returns:
        An openpyxl Workbook ready to save.
    """
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name in SHEET_NAMES:
        df = data.get(sheet_name, pd.DataFrame())
        ws = wb.create_sheet(title=sheet_name)

        if df.empty:
            ws.cell(row=1, column=1, value="No data available")
            ws.column_dimensions["A"].width = 25
            continue

        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        _style_sheet(ws, df)
        _format_currency_columns(ws, df)

    return wb


def save_workbook(wb: Workbook) -> Path:
    """Save the workbook to the outputs directory with a timestamped name.

    Args:
        wb: The openpyxl Workbook to save.

    Returns:
        The absolute path to the saved file.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = OUTPUT_DIR / f"retail_analytics_{timestamp}.xlsx"
    wb.save(str(filepath))
    logger.info("Workbook saved to %s", filepath)
    return filepath


def main() -> None:
    """Main entry point: connect to DB, fetch data, build workbook, save."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    )
    logger.info("Starting RetailFlow Analytics Excel export...")

    try:
        engine = get_engine()
        data = fetch_data(engine)
        wb = build_workbook(data)
        filepath = save_workbook(wb)
        logger.info("Export complete! File: %s", filepath)
        print(f"\nAnalytics export saved to: {filepath}\n")
    except Exception as exc:
        logger.error("Export failed: %s", exc, exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
